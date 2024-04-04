import base64
import html
import io
import json
import logging
import os
import smtplib
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import logging.handlers
import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from dotenv import load_dotenv
load_dotenv()

# Define your log file path here 
log_file_path = "log.txt"

# Create logger
logger = logging.getLogger('my_logger') 
logger.setLevel(logging.INFO) 

# Create console handler and set level to info
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)

# Create file handler and set level to info
file_handler = logging.handlers.RotatingFileHandler(log_file_path, backupCount=3)
file_handler.setLevel(logging.INFO)

# Create formatter
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s', datefmt='%d/%m/%Y %H:%M:%S')

# Add formatter to console handler and file handler
console_handler.setFormatter(formatter)
file_handler.setFormatter(formatter)

# Add console handler and file handler to logger
logger.addHandler(console_handler)
logger.addHandler(file_handler)

smtp_server = "smtp.gmail.com"
smtp_port = 587
smtp_username = os.environ["SMTP_SENDER"]
smtp_password = os.environ["SMTP_PSW"]
CONSUMER_KEY = os.environ["KEY"]
CONSUMER_SECRET = os.environ["SECRET"]
recipient_emails = os.environ['SMTP_TO']
orders_url = os.environ['ORDERS_URL']


auth = base64.b64encode(f"{CONSUMER_KEY}:{CONSUMER_SECRET}".encode("utf-8")).decode("utf-8")

headers = {
		"Content-Type": "application/json",
		"Authorization": f"Basic {auth}"
}

# Define the path to the JSON file
json_file = "processed_orders.json"

try:
		with open(json_file, "r") as f:
				processed_orders = set(json.load(f))
				logger.info(f"Read {len(processed_orders)} existing orders previously processed")
except FileNotFoundError:
		processed_orders = set()
		logger.info("No existing processed orders found")

# Create an empty list to store all the orders
all_orders = []

# Page number
page = 1

while True:
		response = requests.get(f"{orders_url}?per_page=100&page={page}", headers=headers)
		if response.json():
				orders = response.json()
				all_orders.extend(orders)
				logger.info(f"Read {len(orders)} orders from WooCommerce in 'processing' state (page {page})")
				page += 1
		# If response is empty, break the loop.
		else:
				break


new_orders = [
		order for order in all_orders if order["status"] == "processing" and order["id"] not in processed_orders
]
logger.info(f"{len(new_orders)} orders have not been previously processed and will be processed now")

for order in new_orders:
		logger.info(f"Processing order ID: {order['id']}")

		# Create a new Excel workbook
		workbook = Workbook()
		worksheet = workbook.active

		# Add order details to the worksheet
		worksheet["A1"] = "Order ID"
		worksheet["B1"] = order["id"]
		worksheet["A2"] = "Order Date"
		order_date = datetime.strptime(order["date_created"], "%Y-%m-%dT%H:%M:%S")
		worksheet["B2"] = order_date.strftime("%Y-%m-%d %I:%M %p")
		worksheet["A3"] = "Customer Name"
		worksheet["B3"] = f"{html.unescape(order['shipping']['first_name'])} {html.unescape(order['shipping']['last_name'])}"

		# Add shipping details to the worksheet
		worksheet["A4"] = "Shipping Address"
		worksheet["B4"] = f"{html.unescape(order['shipping']['address_1'])}, {html.unescape(order['shipping']['city'])}, {html.unescape(order['shipping']['state'])} {html.unescape(order['shipping']['postcode'])}, {html.unescape(order['shipping']['country'])}"

		# Add payment details to the worksheet
		worksheet["A5"] = "Payment Method"
		worksheet["B5"] = html.unescape(order["payment_method_title"])

		# Add order items to the worksheet
		worksheet["A7"] = "Items"
		worksheet["A7"].font = Font(bold=True)
		row = 8
		for idx, line_item in enumerate(order["line_items"], start=1):
				worksheet.cell(row=row, column=1, value=idx)
				worksheet.cell(row=row, column=2, value=html.unescape(line_item["name"]))
				worksheet.cell(row=row, column=3, value=line_item["quantity"])
				row += 1
				for meta in line_item["meta_data"]:
						if meta["display_key"] and meta["display_value"] and meta["display_key"] != "_wapf_meta" and "(+$" not in meta["display_value"]:
								worksheet.cell(row=row, column=2, value=f"   {html.unescape(meta['display_key'])}")
								worksheet.cell(row=row, column=3, value=html.unescape(meta["display_value"]))
								row += 1

		# Add customer note to the worksheet if available
		if order["customer_note"]:
				worksheet.cell(row=row+1, column=1, value="Customer Note")
				worksheet.cell(row=row+1, column=2, value=html.unescape(order["customer_note"]))

		# Save the Excel file to a BytesIO object
		excel_file = io.BytesIO()
		workbook.save(excel_file)
		excel_file.seek(0)

		# Prepare the packing slip contents
		packing_slip = "Customer Details:\n"
		packing_slip += f"Name: {html.unescape(order['shipping']['first_name'])} {html.unescape(order['shipping']['last_name'])}\n"
		packing_slip += f"Shipping Address: {html.unescape(order['shipping']['address_1'])}, {html.unescape(order['shipping']['city'])}, {html.unescape(order['shipping']['state'])} {html.unescape(order['shipping']['postcode'])}, {html.unescape(order['shipping']['country'])}\n\n"
		packing_slip += "Order Details:\n"
		packing_slip += f"Order Number: {order['id']}\n"
		packing_slip += f"Order Date: {order_date.strftime('%Y-%m-%d %I:%M %p')}\n"
		packing_slip += f"Payment Method: {html.unescape(order['payment_method_title'])}\n\n"
		packing_slip += "Items:\n"

		for idx, line_item in enumerate(order["line_items"], start=1):
				packing_slip += f"\n{idx}. {html.unescape(line_item['name'])} x {line_item['quantity']}\n"
				for meta in line_item["meta_data"]:
						if meta["display_key"] and meta["display_value"] and meta["display_key"] != "_wapf_meta":
								display_value = meta["display_value"]
								if "(+$" in display_value:
										display_value = display_value.split("(+$")[0].strip()
								packing_slip += f"   {html.unescape(meta['display_key'])}: {html.unescape(display_value)}\n"

		if order["customer_note"]:
				packing_slip += f"\nCustomer Note: {html.unescape(order['customer_note'])}\n"

		packing_slip += "\n--------------------"

		# Create the email message with multipart support
		msg = MIMEMultipart()
		msg["Subject"] = f"New Order - Order ID {order['id']}"
		msg["From"] = smtp_username
		msg['To'] = recipient_emails

		# Attach the packing slip as text
		text = MIMEText(packing_slip)
		msg.attach(text)

		# Attach the Excel spreadsheet
		part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
		part.set_payload(excel_file.read())
		encoders.encode_base64(part)
		part.add_header("Content-Disposition", f"attachment; filename=order_{order['id']}.xlsx",)
		msg.attach(part)

		try:
				with smtplib.SMTP(smtp_server, smtp_port) as server:
						server.starttls()
						server.login(smtp_username, smtp_password)
						server.send_message(msg)
						logger.info(f"Email sent for order ID {order['id']}")
		except smtplib.SMTPException as e:
				logger.error(f"Error sending email for order ID {order['id']}: {e}")

		# Print the packing slip contents to the console
		print(packing_slip)

		# Add the order ID to the set of processed orders
		processed_orders.add(order['id'])

# Save the processed orders to the JSON file
with open(json_file, 'w') as f:
		json.dump(list(processed_orders), f)
		logger.info(f"Saved {len(processed_orders)} processed orders to the JSON file")